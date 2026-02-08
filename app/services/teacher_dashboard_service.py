"""Service for calculating teacher dashboard statistics (class-level).

This module provides comprehensive class-level analytics for teachers, including
student progress tracking, performance metrics, and data export functionality.

Autor: Gernibide
"""

import csv
import io
import time
from collections.abc import Callable
from datetime import datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from app.models.actividad import Actividad
from app.models.actividad_progreso import ActividadProgreso
from app.models.clase import Clase
from app.models.juego import Partida
from app.models.punto import Punto
from app.models.usuario import Usuario


class CacheEntry:
    """Cache entry with TTL (Time To Live) for temporary data storage.

    Attributes:
        data: The cached data of any type.
        expires_at: Unix timestamp when this cache entry expires.
    """

    def __init__(self, data: Any, ttl_seconds: int):
        """Initialize cache entry.

        Args:
            data: The data to cache.
            ttl_seconds: Time to live in seconds.
        """
        self.data = data
        self.expires_at = time.time() + ttl_seconds

    def is_expired(self) -> bool:
        """Check if cache entry has expired.

        Returns:
            True if current time exceeds expiration time, False otherwise.
        """
        return time.time() > self.expires_at


class TeacherDashboardService:
    """Service for calculating teacher dashboard statistics with caching.

    This service provides comprehensive class-level metrics for teachers including:
    - Class summary statistics (students, progress, time, grades)
    - Individual student progress and performance tracking
    - Activity completion status by class
    - Class evolution over time
    - Data export to CSV and Excel formats

    All methods utilize an in-memory cache with configurable TTL to reduce
    database load. Teacher-specific data has a shorter TTL (2 minutes) for
    more current information.

    Attributes:
        _cache: Class-level cache storage for computed statistics.
        CACHE_TTL: Default cache time-to-live in seconds (120 = 2 minutes).
    """

    # Cache storage
    _cache: dict[str, CacheEntry] = {}

    # Cache TTL in seconds (2 minutes for teacher dashboard)
    CACHE_TTL = 120

    @classmethod
    def _get_cached_or_fetch(
        cls, cache_key: str, fetch_func: Callable, *args, ttl: int | None = None
    ) -> Any:
        """Generic cache getter with TTL.

        Args:
            cache_key: Unique key for this cached data.
            fetch_func: Function to call if cache miss.
            *args: Arguments to pass to fetch_func.
            ttl: Custom TTL in seconds. Uses CACHE_TTL if None.

        Returns:
            Cached or freshly fetched data.
        """
        # Check cache
        if cache_key in cls._cache:
            entry = cls._cache[cache_key]
            if not entry.is_expired():
                return entry.data

        # Cache miss or expired - fetch new data
        data = fetch_func(*args)

        # Store in cache
        ttl_seconds = ttl if ttl is not None else cls.CACHE_TTL
        cls._cache[cache_key] = CacheEntry(data, ttl_seconds)

        return data

    @classmethod
    def clear_cache(cls):
        """Clear all cached data.

        Use this method when you need to force refresh of all statistics,
        for example after bulk data imports or updates.
        """
        cls._cache.clear()

    @staticmethod
    def get_class_summary(
        db: Session, profesor_id: str, clase_id: str | None = None, days: int = 7
    ) -> dict[str, Any]:
        """Get summary statistics for a class (with caching).

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose class(es) to analyze.
            clase_id: Optional specific class ID. If None, aggregates all classes.
            days: Number of days to look back for time/grade metrics. Defaults to 7.

        Returns:
            Dictionary containing:
                - total_alumnos: Number of students in the class(es)
                - progreso_medio: Average progress percentage (0-100)
                - tiempo_promedio: Average time spent in minutes
                - nota_media: Average grade/score
                - clase_nombre: Name of the class or "Todas las clases"
        """
        cache_key = f"class_summary_{profesor_id}_{clase_id}_{days}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key, TeacherDashboardService._fetch_class_summary, db, profesor_id, clase_id, days
        )

    @staticmethod
    def _fetch_class_summary(
        db: Session, profesor_id: str, clase_id: str | None, days: int
    ) -> dict[str, Any]:
        """Internal method to fetch class summary from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.
            clase_id: Optional specific class ID.
            days: Number of days to look back.

        Returns:
            Dictionary with class summary metrics.
        """
        # Build base query for students
        students_query = (
            db.query(Usuario)
            .join(Clase, Usuario.id_clase == Clase.id)
            .filter(Clase.id_profesor == profesor_id)
        )

        if clase_id:
            students_query = students_query.filter(Usuario.id_clase == clase_id)

        students = students_query.all()
        total_students = len(students)

        if total_students == 0:
            return {
                "total_alumnos": 0,
                "progreso_medio": 0,
                "tiempo_promedio": 0,
                "nota_media": 0,
                "clase_nombre": "Sin clase",
            }

        student_ids = [s.id for s in students]

        # Calculate average progress (based on completed activities)
        # Progress = (completed events / total unique events per student) * 100
        total_events = db.query(func.count(func.distinct(Actividad.id))).scalar() or 1

        completed_events_per_student = []
        for student_id in student_ids:
            completed = (
                db.query(func.count(func.distinct(ActividadProgreso.id_actividad)))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(Partida.id_usuario == student_id)
                        ),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )
            completed_events_per_student.append(completed)

        avg_completed = (
            sum(completed_events_per_student) / len(completed_events_per_student)
            if completed_events_per_student
            else 0
        )
        progreso_medio = min(100, (avg_completed / total_events) * 100)

        # Calculate average time from activity durations (duration in minutes)
        fecha_limite = datetime.now() - timedelta(days=days)

        # Sum all activity durations per student, then average across students
        total_time_per_student = []
        for student_id in student_ids:
            student_time = (
                db.query(func.sum(ActividadProgreso.duracion))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(
                                and_(
                                    Partida.id_usuario == student_id,
                                    Partida.fecha_inicio >= fecha_limite,
                                )
                            )
                        ),
                        ActividadProgreso.duracion.isnot(None),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )
            if student_time > 0:
                total_time_per_student.append(student_time)

        avg_time = (
            sum(total_time_per_student) / len(total_time_per_student)
            if total_time_per_student
            else 0
        )

        tiempo_promedio = round(avg_time / 60, 0) if avg_time else 0

        # Calculate average grade (from completed activities with scores)
        avg_grade = (
            db.query(func.avg(ActividadProgreso.puntuacion))
            .filter(
                and_(
                    ActividadProgreso.id_juego.in_(
                        db.query(Partida.id).filter(
                            and_(
                                Partida.id_usuario.in_(student_ids),
                                Partida.fecha_inicio >= fecha_limite,
                            )
                        )
                    ),
                    ActividadProgreso.puntuacion.isnot(None),
                    ActividadProgreso.estado == "completado",
                )
            )
            .scalar()
            or 0
        )

        # Get class name
        if clase_id:
            clase = db.query(Clase).filter(Clase.id == clase_id).first()
            clase_nombre = clase.nombre if clase else "Sin clase"
        else:
            clase_nombre = "Todas las clases"

        return {
            "total_alumnos": total_students,
            "progreso_medio": round(progreso_medio, 1),
            "tiempo_promedio": int(tiempo_promedio),
            "nota_media": round(avg_grade, 1),
            "clase_nombre": clase_nombre,
        }

    @staticmethod
    def get_student_progress(
        db: Session, profesor_id: str, clase_id: str | None = None
    ) -> dict[str, list]:
        """Get progress for each student in the class (with caching).

        Calculates the percentage of unique activities completed by each student.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose class(es) to analyze.
            clase_id: Optional specific class ID. If None, includes all classes.

        Returns:
            Dictionary containing:
                - students: List of student full names
                - progress: Progress percentage (0-100) for each student
        """
        cache_key = f"student_progress_{profesor_id}_{clase_id}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key, TeacherDashboardService._fetch_student_progress, db, profesor_id, clase_id
        )

    @staticmethod
    def _fetch_student_progress(
        db: Session, profesor_id: str, clase_id: str | None
    ) -> dict[str, list]:
        """Internal method to fetch student progress from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.
            clase_id: Optional specific class ID.

        Returns:
            Dictionary with student progress data.
        """
        # Get students
        students_query = (
            db.query(Usuario)
            .join(Clase, Usuario.id_clase == Clase.id)
            .filter(Clase.id_profesor == profesor_id)
        )

        if clase_id:
            students_query = students_query.filter(Usuario.id_clase == clase_id)

        students = students_query.all()

        if not students:
            return {"students": [], "progress": []}

        total_events = db.query(func.count(func.distinct(Actividad.id))).scalar() or 1

        student_names = []
        progress_values = []

        for student in students:
            # Calculate progress for this student
            completed = (
                db.query(func.count(func.distinct(ActividadProgreso.id_actividad)))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(Partida.id_usuario == student.id)
                        ),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )

            progress = min(100, (completed / total_events) * 100)

            student_names.append(f"{student.nombre} {student.apellido}")
            progress_values.append(round(progress, 1))

        return {"students": student_names, "progress": progress_values}

    @staticmethod
    def get_student_time(
        db: Session, profesor_id: str, clase_id: str | None = None, days: int = 7
    ) -> dict[str, list]:
        """Get time spent for each student in the class (with caching).

        Calculates total time spent on completed activities within the specified
        time window. Only includes students with time > 0.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose class(es) to analyze.
            clase_id: Optional specific class ID. If None, includes all classes.
            days: Number of days to look back. Defaults to 7.

        Returns:
            Dictionary containing:
                - students: List of student first names (students with time > 0 only)
                - time: Time spent in minutes for each student
        """
        cache_key = f"student_time_{profesor_id}_{clase_id}_{days}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key, TeacherDashboardService._fetch_student_time, db, profesor_id, clase_id, days
        )

    @staticmethod
    def _fetch_student_time(
        db: Session, profesor_id: str, clase_id: str | None, days: int
    ) -> dict[str, list]:
        """Internal method to fetch student time from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.
            clase_id: Optional specific class ID.
            days: Number of days to look back.

        Returns:
            Dictionary with student time data.
        """
        # Get students
        students_query = (
            db.query(Usuario)
            .join(Clase, Usuario.id_clase == Clase.id)
            .filter(Clase.id_profesor == profesor_id)
        )

        if clase_id:
            students_query = students_query.filter(Usuario.id_clase == clase_id)

        students = students_query.all()

        if not students:
            return {"students": [], "time": []}

        fecha_limite = datetime.now() - timedelta(days=days)

        student_names = []
        time_values = []

        for student in students:
            # Calculate total time for this student from activity durations
            total_time = (
                db.query(func.sum(ActividadProgreso.duracion))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(
                                and_(
                                    Partida.id_usuario == student.id,
                                    Partida.fecha_inicio >= fecha_limite,
                                )
                            )
                        ),
                        ActividadProgreso.duracion.isnot(None),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )

            time_minutes = round(total_time / 60, 0) if total_time else 0

            # Only include students with time > 0
            if time_minutes > 0:
                student_names.append(student.nombre)  # First name only for chart
                time_values.append(int(time_minutes))

        return {"students": student_names, "time": time_values}

    @staticmethod
    def get_activities_by_class(
        db: Session, profesor_id: str, clase_id: str | None = None
    ) -> dict[str, Any]:
        """Get activity completion status by class (with caching).

        Analyzes how many students have completed, started, or not started
        each activity punto.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose class(es) to analyze.
            clase_id: Optional specific class ID. If None, includes all classes.

        Returns:
            Dictionary containing:
                - activities: List of activity punto names
                - completed: Number of students who completed each activity
                - in_progress: Number of students with activity in progress
                - not_started: Number of students who haven't started
        """
        cache_key = f"activities_by_class_{profesor_id}_{clase_id}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key, TeacherDashboardService._fetch_activities_by_class, db, profesor_id, clase_id
        )

    @staticmethod
    def _fetch_activities_by_class(
        db: Session, profesor_id: str, clase_id: str | None
    ) -> dict[str, Any]:
        """Internal method to fetch activities by class from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.
            clase_id: Optional specific class ID.

        Returns:
            Dictionary with activity completion status data.
        """
        # Get students
        students_query = (
            db.query(Usuario)
            .join(Clase, Usuario.id_clase == Clase.id)
            .filter(Clase.id_profesor == profesor_id)
        )

        if clase_id:
            students_query = students_query.filter(Usuario.id_clase == clase_id)

        students = students_query.all()
        student_ids = [s.id for s in students]
        total_students = len(student_ids)

        if total_students == 0:
            return {"activities": [], "completed": [], "in_progress": [], "not_started": []}

        # Get all activities
        activities = db.query(Punto).all()

        activity_names = []
        completed_counts = []
        in_progress_counts = []
        not_started_counts = []

        for activity in activities:
            activity_names.append(activity.nombre)

            # Count students who completed this activity
            completed = (
                db.query(func.count(func.distinct(Partida.id_usuario)))
                .filter(
                    and_(
                        Partida.id_usuario.in_(student_ids),
                        Partida.id.in_(
                            db.query(ActividadProgreso.id_juego).filter(
                                and_(
                                    ActividadProgreso.id_punto == activity.id,
                                    ActividadProgreso.estado == "completado",
                                )
                            )
                        ),
                    )
                )
                .scalar()
                or 0
            )

            # Count students who have this activity in progress
            in_progress = (
                db.query(func.count(func.distinct(Partida.id_usuario)))
                .filter(
                    and_(
                        Partida.id_usuario.in_(student_ids),
                        Partida.id.in_(
                            db.query(ActividadProgreso.id_juego).filter(
                                and_(
                                    ActividadProgreso.id_punto == activity.id,
                                    ActividadProgreso.estado == "en_progreso",
                                )
                            )
                        ),
                        # Exclude if already completed
                        ~Partida.id.in_(
                            db.query(ActividadProgreso.id_juego).filter(
                                and_(
                                    ActividadProgreso.id_punto == activity.id,
                                    ActividadProgreso.estado == "completado",
                                )
                            )
                        ),
                    )
                )
                .scalar()
                or 0
            )

            # Not started = total - (completed + in_progress)
            not_started = max(0, total_students - completed - in_progress)

            completed_counts.append(completed)
            in_progress_counts.append(in_progress)
            not_started_counts.append(not_started)

        return {
            "activities": activity_names,
            "completed": completed_counts,
            "in_progress": in_progress_counts,
            "not_started": not_started_counts,
        }

    @staticmethod
    def get_class_evolution(
        db: Session, profesor_id: str, clase_id: str | None = None, days: int = 14
    ) -> dict[str, Any]:
        """Get class evolution over time (with caching).

        Tracks how the class's average progress and grades have evolved
        over the specified time period.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose class(es) to analyze.
            clase_id: Optional specific class ID. If None, includes all classes.
            days: Number of days to retrieve. Defaults to 14.

        Returns:
            Dictionary containing:
                - dates: List of date strings (YYYY-MM-DD format)
                - progress: Average class progress percentage for each date
                - grades: Average class grade for each date
        """
        cache_key = f"class_evolution_{profesor_id}_{clase_id}_{days}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key,
            TeacherDashboardService._fetch_class_evolution,
            db,
            profesor_id,
            clase_id,
            days,
        )

    @staticmethod
    def _fetch_class_evolution(
        db: Session, profesor_id: str, clase_id: str | None, days: int
    ) -> dict[str, Any]:
        """Internal method to fetch class evolution from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.
            clase_id: Optional specific class ID.
            days: Number of days to retrieve.

        Returns:
            Dictionary with class evolution data.
        """
        # Get students
        students_query = (
            db.query(Usuario)
            .join(Clase, Usuario.id_clase == Clase.id)
            .filter(Clase.id_profesor == profesor_id)
        )

        if clase_id:
            students_query = students_query.filter(Usuario.id_clase == clase_id)

        students = students_query.all()
        student_ids = [s.id for s in students]

        if not student_ids:
            return {"dates": [], "progress": [], "grades": []}

        total_events = db.query(func.count(func.distinct(Actividad.id))).scalar() or 1

        dates = []
        progress_values = []
        grade_values = []

        now = datetime.now()

        for i in range(days - 1, -1, -1):
            date = now - timedelta(days=i)
            date_start = datetime(date.year, date.month, date.day)
            date_end = date_start + timedelta(days=1)

            # Calculate average progress for this date
            completed_by_date = []
            for student_id in student_ids:
                completed = (
                    db.query(func.count(func.distinct(ActividadProgreso.id_actividad)))
                    .filter(
                        and_(
                            ActividadProgreso.id_juego.in_(
                                db.query(Partida.id).filter(
                                    and_(
                                        Partida.id_usuario == student_id,
                                        Partida.fecha_inicio < date_end,
                                    )
                                )
                            ),
                            ActividadProgreso.estado == "completado",
                            ActividadProgreso.fecha_inicio < date_end,
                        )
                    )
                    .scalar()
                    or 0
                )
                completed_by_date.append(completed)

            avg_completed = (
                sum(completed_by_date) / len(completed_by_date) if completed_by_date else 0
            )
            progress = min(100, (avg_completed / total_events) * 100)

            # Calculate average grade for this date
            avg_grade = (
                db.query(func.avg(ActividadProgreso.puntuacion))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(
                                and_(
                                    Partida.id_usuario.in_(student_ids),
                                    Partida.fecha_inicio < date_end,
                                )
                            )
                        ),
                        ActividadProgreso.puntuacion.isnot(None),
                        ActividadProgreso.estado == "completado",
                        ActividadProgreso.fecha_inicio < date_end,
                    )
                )
                .scalar()
                or 0
            )

            dates.append(date_start.strftime("%Y-%m-%d"))
            progress_values.append(round(progress, 1))
            grade_values.append(round(avg_grade, 1) if avg_grade else 0)

        return {"dates": dates, "progress": progress_values, "grades": grade_values}

    @staticmethod
    def get_profesor_classes(db: Session, profesor_id: str) -> list[dict[str, str]]:
        """Get all classes for a professor (with caching).

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.

        Returns:
            List of dictionaries, each containing:
                - id: Class ID
                - nombre: Class name
        """
        cache_key = f"profesor_classes_{profesor_id}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key,
            TeacherDashboardService._fetch_profesor_classes,
            db,
            profesor_id,
            ttl=600,  # Cache for 10 minutes
        )

    @staticmethod
    def _fetch_profesor_classes(db: Session, profesor_id: str) -> list[dict[str, str]]:
        """Internal method to fetch professor classes from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.

        Returns:
            List of class dictionaries.
        """
        classes = db.query(Clase).filter(Clase.id_profesor == profesor_id).all()

        return [{"id": clase.id, "nombre": clase.nombre} for clase in classes]

    @staticmethod
    def get_students_list(
        db: Session, profesor_id: str, clase_id: str | None = None
    ) -> list[dict[str, Any]]:
        """Get detailed list of students with progress and performance data (with caching).

        Provides comprehensive information for each student including progress,
        time spent, grades, and last activity date. Results are sorted by
        progress descending.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose students to retrieve.
            clase_id: Optional specific class ID. If None, includes all classes.

        Returns:
            List of student dictionaries, each containing:
                - nombre: Full name (first + last)
                - username: Student username
                - progreso: Progress percentage (0-100)
                - tiempo_total: Total time spent in minutes
                - nota_media: Average grade/score
                - actividades_completadas: Number of unique activities completed
                - ultima_actividad: Last activity date (YYYY-MM-DD) or "Nunca"
        """
        cache_key = f"students_list_{profesor_id}_{clase_id}"
        return TeacherDashboardService._get_cached_or_fetch(
            cache_key,
            TeacherDashboardService._fetch_students_list,
            db,
            profesor_id,
            clase_id,
            ttl=60,  # Cache for 1 minute
        )

    @staticmethod
    def _fetch_students_list(
        db: Session, profesor_id: str, clase_id: str | None
    ) -> list[dict[str, Any]]:
        """Internal method to fetch students list from database.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor.
            clase_id: Optional specific class ID.

        Returns:
            List of student detail dictionaries.
        """
        # Get students
        students_query = (
            db.query(Usuario)
            .join(Clase, Usuario.id_clase == Clase.id)
            .filter(Clase.id_profesor == profesor_id)
        )

        if clase_id:
            students_query = students_query.filter(Usuario.id_clase == clase_id)

        students = students_query.all()

        if not students:
            return []

        total_activities = db.query(func.count(func.distinct(Actividad.id))).scalar() or 1

        students_data = []

        for student in students:
            # Calculate completed activities
            completed_activities = (
                db.query(func.count(func.distinct(ActividadProgreso.id_actividad)))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(Partida.id_usuario == student.id)
                        ),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )

            # Calculate progress percentage
            progreso = min(100, (completed_activities / total_activities) * 100)

            # Calculate total time from sum of activity durations (not partida duration)
            total_time = (
                db.query(func.sum(ActividadProgreso.duracion))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(Partida.id_usuario == student.id)
                        ),
                        ActividadProgreso.duracion.isnot(None),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )

            tiempo_minutos = round(total_time / 60, 0) if total_time else 0

            # Calculate average grade
            avg_grade = (
                db.query(func.avg(ActividadProgreso.puntuacion))
                .filter(
                    and_(
                        ActividadProgreso.id_juego.in_(
                            db.query(Partida.id).filter(Partida.id_usuario == student.id)
                        ),
                        ActividadProgreso.puntuacion.isnot(None),
                        ActividadProgreso.estado == "completado",
                    )
                )
                .scalar()
                or 0
            )

            # Get last activity date
            last_activity = (
                db.query(func.max(Partida.fecha_inicio))
                .filter(Partida.id_usuario == student.id)
                .scalar()
            )

            students_data.append(
                {
                    "nombre": f"{student.nombre} {student.apellido}",
                    "username": student.username,
                    "progreso": round(progreso, 1),
                    "tiempo_total": int(tiempo_minutos),
                    "nota_media": round(avg_grade, 1) if avg_grade else 0,
                    "actividades_completadas": completed_activities,
                    "ultima_actividad": (
                        last_activity.strftime("%Y-%m-%d") if last_activity else "Nunca"
                    ),
                }
            )

        # Sort by progress descending
        students_data.sort(key=lambda x: x["progreso"], reverse=True)

        return students_data

    @staticmethod
    def export_students_csv(db: Session, profesor_id: str, clase_id: str | None = None) -> str:
        """Export students list to CSV format.

        Generates a CSV file with student performance data. Uses the cached
        students list from get_students_list.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose students to export.
            clase_id: Optional specific class ID. If None, includes all classes.

        Returns:
            CSV content as string, ready for download.
        """
        students_data = TeacherDashboardService.get_students_list(db, profesor_id, clase_id)

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(
            [
                "Nombre",
                "Usuario",
                "Progreso (%)",
                "Actividades Completadas",
                "Tiempo Total (min)",
                "Nota Media",
                "Última Actividad",
            ]
        )

        # Write data rows
        for student in students_data:
            writer.writerow(
                [
                    student["nombre"],
                    student["username"],
                    student["progreso"],
                    student["actividades_completadas"],
                    student["tiempo_total"],
                    student["nota_media"],
                    student["ultima_actividad"],
                ]
            )

        return output.getvalue()

    @staticmethod
    def export_students_excel(db: Session, profesor_id: str, clase_id: str | None = None) -> bytes:
        """Export students list to Excel format.

        Generates a formatted Excel workbook with student performance data,
        including header styling and proper column widths. Uses the cached
        students list from get_students_list.

        Args:
            db: Database session for querying.
            profesor_id: ID of the professor whose students to export.
            clase_id: Optional specific class ID. If None, includes all classes.

        Returns:
            Excel file as bytes, ready for download.
        """
        students_data = TeacherDashboardService.get_students_list(db, profesor_id, clase_id)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alumnos"

        # Define styles
        header_fill = PatternFill(start_color="6B8E3A", end_color="6B8E3A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = [
            "Nombre",
            "Usuario",
            "Progreso (%)",
            "Actividades Completadas",
            "Tiempo Total (min)",
            "Nota Media",
            "Última Actividad",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_num, student in enumerate(students_data, 2):
            ws.cell(row=row_num, column=1, value=student["nombre"])
            ws.cell(row=row_num, column=2, value=student["username"])
            ws.cell(row=row_num, column=3, value=student["progreso"])
            ws.cell(row=row_num, column=4, value=student["actividades_completadas"])
            ws.cell(row=row_num, column=5, value=student["tiempo_total"])
            ws.cell(row=row_num, column=6, value=student["nota_media"])
            ws.cell(row=row_num, column=7, value=student["ultima_actividad"])

        # Adjust column widths
        column_widths = [30, 20, 15, 25, 20, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()
